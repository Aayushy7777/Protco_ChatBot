#!/usr/bin/env python3
"""
Test script to validate Excel file upload support (XLS and XLSX).
"""
import requests
import json
import io
from openpyxl import Workbook
import xlwt

BASE_URL = "http://localhost:8000"

def create_test_xlsx():
    """Create a test XLSX file."""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Name'
    ws['B1'] = 'Amount'
    ws['C1'] = 'Date'
    ws['A2'] = 'Item A'
    ws['B2'] = 1000
    ws['C2'] = '2024-01-15'
    ws['A3'] = 'Item B'
    ws['B3'] = 2500
    ws['C3'] = '2024-02-20'
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

def create_test_xls():
    """Create a test XLS file."""
    wb = xlwt.Workbook()
    ws = wb.add_sheet('Sheet1')
    ws.write(0, 0, 'Name')
    ws.write(0, 1, 'Amount')
    ws.write(0, 2, 'Date')
    ws.write(1, 0, 'Invoice A')
    ws.write(1, 1, 500)
    ws.write(1, 2, '2024-01-01')
    ws.write(2, 0, 'Invoice B')
    ws.write(2, 1, 1500)
    ws.write(2, 2, '2024-02-15')
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

def test_upload(filename, content, file_type):
    """Test uploading a file."""
    files = {'files': (filename, io.BytesIO(content), file_type)}
    response = requests.post(f"{BASE_URL}/api/upload", files=files)
    return response.status_code, response.json()

if __name__ == "__main__":
    print("Testing Excel file uploads...\n")
    
    # Test XLSX
    print("Testing XLSX upload...")
    xlsx_content = create_test_xlsx()
    status, response = test_upload("test_invoice.xlsx", xlsx_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    print(f"  Status: {status}")
    if status == 200:
        upload_info = response.get('uploaded', [{}])[0]
        print(f"  ✓ Rows: {upload_info.get('rows')}, Columns: {upload_info.get('columns')}")
    else:
        print(f"  ✗ Error: {response.get('errors', response)}")
    
    # Test XLS
    print("\nTesting XLS upload...")
    xls_content = create_test_xls()
    status, response = test_upload("test_report.xls", xls_content, "application/vnd.ms-excel")
    print(f"  Status: {status}")
    if status == 200:
        upload_info = response.get('uploaded', [{}])[0]
        print(f"  ✓ Rows: {upload_info.get('rows')}, Columns: {upload_info.get('columns')}")
    else:
        print(f"  ✗ Error: {response.get('errors', response)}")
    
    print("\n✓ Excel file upload test complete!")
